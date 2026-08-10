import pandas as pd
import warnings
from datetime import date
from typing import Optional
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from format_config import (
    SalaryColumns,
    ScheduleColumns,
    month_column,
    format_period,
    parse_period,
)

from scheduler import WorkDateScheduler


class DaycareAllocator:
    def __init__(self, hourly_rate: float, hours_per_daycare: int = 1, scheduler: Optional[WorkDateScheduler] = None):
        """Create allocator.

        Pass a custom `scheduler` for different work-day rules during tests
        or if you want to tweak holiday/eve rules without changing code.
        """
        self.hourly_rate = hourly_rate
        self.hours_per_daycare = self._normalize_hours_per_daycare(hours_per_daycare)
        self.scheduler = scheduler or WorkDateScheduler()

    def _normalize_hours_per_daycare(self, hours_per_daycare: int) -> int:
        try:
            hours = int(hours_per_daycare)
        except (TypeError, ValueError) as exc:
            raise ValueError("כמות השעות למעון חייבת להיות מספר שלם חיובי.") from exc

        if hours <= 0:
            raise ValueError("כמות השעות למעון חייבת להיות מספר שלם חיובי.")
        return hours

    def _read_salaries(self, filepath: str) -> pd.DataFrame:
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            try:
                df = pd.read_csv(filepath, encoding='utf-8', index_col=False)
            except UnicodeDecodeError:
                df = pd.read_csv(filepath, encoding='windows-1255', index_col=False)
        
        df.columns = df.columns.str.strip()
        
        if SalaryColumns.FIRST_NAME not in df.columns or SalaryColumns.LAST_NAME not in df.columns:
            raise ValueError(
                f"עמודות '{SalaryColumns.FIRST_NAME}' או '{SalaryColumns.LAST_NAME}' חסרות בקובץ השכר."
            )
            
        df[SalaryColumns.DISPLAY_NAME] = (
            df[SalaryColumns.FIRST_NAME].astype(str).str.strip() + " " +
            df[SalaryColumns.LAST_NAME].astype(str).str.strip()
        )
        return df

    def _extract_work_months(self, df_salaries: pd.DataFrame) -> list:
        if SalaryColumns.PERIOD not in df_salaries.columns:
            raise ValueError(f"עמודת '{SalaryColumns.PERIOD}' לא נמצאה בקובץ השכר.")
            
        periods = df_salaries[SalaryColumns.PERIOD].dropna().astype(str).str.strip().unique()
        months_list = []
        for p in periods:
            parsed = parse_period(p)
            if parsed is not None:
                months_list.append(parsed)
        return sorted(months_list)

    def _get_worker_capacity(self, df_salaries: pd.DataFrame, display_name: str, year: int, month: int) -> int:
        period_str = format_period(year, month)
        worker_data = df_salaries[
            (df_salaries[SalaryColumns.DISPLAY_NAME] == display_name) & 
            (df_salaries[SalaryColumns.PERIOD].astype(str).str.strip() == period_str)
        ]
        
        if worker_data.empty:
            return 0
            
        salary = float(worker_data[SalaryColumns.TOTAL_EMPLOYER_COST].iloc[0])
        total_hours = int(salary // self.hourly_rate)
        if self.hours_per_daycare <= 0:
            return 0
        return total_hours // self.hours_per_daycare

    def generate_schedule(self, daycares_path: str, salaries_path: str, output_path: str) -> list:
        df_daycares = pd.read_excel(daycares_path)
        daycares_list = df_daycares.iloc[:, 0].dropna().unique().tolist()
        
        daycare_symbols = {}
        if len(df_daycares.columns) > 1:
            for _, row in df_daycares.iterrows():
                daycare_symbols[row.iloc[0]] = row.iloc[1]

        df_salaries = self._read_salaries(salaries_path)
        available_workers = df_salaries[SalaryColumns.DISPLAY_NAME].unique().tolist()
        work_months = self._extract_work_months(df_salaries)
        
        if not available_workers or not work_months:
            raise ValueError("חסרים נתונים מקובץ השכר.")

        # 1. שיוך "מעונות אם" בגושים רציפים לפי סדר ההופעה בקובץ
        primary_worker = {}
        n_daycares = len(daycares_list)
        n_workers = len(available_workers)
        base_chunk = n_daycares // n_workers
        remainder = n_daycares % n_workers

        current_idx = 0
        for i, w in enumerate(available_workers):
            # המדריכות הראשונות יקבלו מעון אחד נוסף אם החלוקה אינה שלמה
            chunk_size = base_chunk + (1 if i < remainder else 0)
            for _ in range(chunk_size):
                primary_worker[daycares_list[current_idx]] = w
                current_idx += 1

        granted_shifts = {}
        unassigned_warnings = []
        
        for year, month in work_months:
            granted_shifts[(year, month)] = {}
            capacities = {w: self._get_worker_capacity(df_salaries, w, year, month) for w in available_workers}
            unassigned_daycares = list(daycares_list) # הרשימה נשמרת לפי סדר ההופעה בקובץ
            
            # שלב א': שיבוץ מדריכות קבועות למעונות האם
            for d in list(unassigned_daycares):
                w = primary_worker[d]
                if capacities[w] > 0:
                    granted_shifts[(year, month)][d] = w
                    capacities[w] -= 1
                    unassigned_daycares.remove(d)
                    
            # שלב ב': העברת מעונות יתומים (לפי סדרם) למדריכות עם שעות פנויות
            for d in list(unassigned_daycares):
                available_spares = {w: cap for w, cap in capacities.items() if cap > 0}
                if available_spares:
                    # בוחר את העובדת עם הכי הרבה שעות פנויות כרגע
                    w_spare = max(available_spares, key=available_spares.get)
                    granted_shifts[(year, month)][d] = w_spare
                    capacities[w_spare] -= 1
                    unassigned_daycares.remove(d)
                    
            # שלב ג': תיעוד מעונות שנותרו ללא פתרון
            for d in unassigned_daycares:
                unassigned_warnings.append(f"חודש {month:02d}/{year}: המעון '{d}'")

        worker_all_daycares = {w: [] for w in available_workers}
        for w in available_workers:
            my_primary = [d for d in daycares_list if primary_worker[d] == w]
            worker_all_daycares[w].extend(my_primary)
            
            my_extra = []
            for year, month in work_months:
                for d, assigned_w in granted_shifts[(year, month)].items():
                    if assigned_w == w and d not in worker_all_daycares[w] and d not in my_extra:
                        my_extra.append(d)
            worker_all_daycares[w].extend(my_extra)

        worker_month_dates = {}
        for w in available_workers:
            for year, month in work_months:
                n_shifts = list(granted_shifts[(year, month)].values()).count(w)
                dates_list = self.scheduler.get_evenly_distributed_dates(year, month, n_shifts)
                worker_month_dates[(w, year, month)] = dates_list

        records = []
        for worker in available_workers:
            daycares_for_w = worker_all_daycares[worker]
            if not daycares_for_w:
                continue
                
            worker_totals_hours = {month_column(ScheduleColumns.HOURS_TEMPLATE, m[1]): 0 for m in work_months}
            worker_totals_pay = {month_column(ScheduleColumns.RATE_TEMPLATE, m[1]): 0 for m in work_months}
            
            for i, daycare in enumerate(daycares_for_w):
                row_data = {
                    ScheduleColumns.WORKER_NAME: worker if i == 0 else "",
                    ScheduleColumns.DAYCARE_NAME: daycare,
                    ScheduleColumns.SYMBOL: daycare_symbols.get(daycare, "")
                }
                
                for year, month in work_months:
                    col_date = month_column(ScheduleColumns.DATE_TEMPLATE, month)
                    col_hours = month_column(ScheduleColumns.HOURS_TEMPLATE, month)
                    col_rate = month_column(ScheduleColumns.RATE_TEMPLATE, month)
                    
                    if granted_shifts[(year, month)].get(daycare) == worker:
                        if worker_month_dates[(worker, year, month)]:
                            w_date = worker_month_dates[(worker, year, month)].pop(0)
                            row_data[col_date] = w_date.strftime('%d/%m/%Y')
                            row_data[col_hours] = self.hours_per_daycare
                            row_data[col_rate] = self.hourly_rate * self.hours_per_daycare
                            worker_totals_hours[col_hours] += self.hours_per_daycare
                            worker_totals_pay[col_rate] += self.hourly_rate * self.hours_per_daycare
                        else:
                            row_data[col_date], row_data[col_hours], row_data[col_rate] = "", "", ""
                    else:
                        row_data[col_date], row_data[col_hours], row_data[col_rate] = "", "", ""
                
                records.append(row_data)

            sum_hours_row = {
                ScheduleColumns.WORKER_NAME: '',
                ScheduleColumns.DAYCARE_NAME: ScheduleColumns.TOTAL_HOURS_LABEL,
                ScheduleColumns.SYMBOL: ''
            }
            sum_pay_row = {
                ScheduleColumns.WORKER_NAME: '',
                ScheduleColumns.DAYCARE_NAME: ScheduleColumns.TOTAL_PAY_LABEL,
                ScheduleColumns.SYMBOL: ''
            }
            
            for year, month in work_months:
                col_date = month_column(ScheduleColumns.DATE_TEMPLATE, month)
                col_hours = month_column(ScheduleColumns.HOURS_TEMPLATE, month)
                col_rate = month_column(ScheduleColumns.RATE_TEMPLATE, month)
                
                sum_hours_row[col_date] = ""
                sum_hours_row[col_hours] = worker_totals_hours[col_hours] if worker_totals_hours[col_hours] > 0 else ""
                sum_hours_row[col_rate] = ""
                
                sum_pay_row[col_date] = ""
                sum_pay_row[col_hours] = ""
                sum_pay_row[col_rate] = worker_totals_pay[col_rate] if worker_totals_pay[col_rate] > 0 else ""

            records.append(sum_hours_row)
            records.append(sum_pay_row)

        schedule_df = pd.DataFrame(records)
        
        cols_order = [
            ScheduleColumns.WORKER_NAME,
            ScheduleColumns.DAYCARE_NAME,
            ScheduleColumns.SYMBOL,
        ]
        for year, month in work_months:
            cols_order.extend([
                month_column(ScheduleColumns.DATE_TEMPLATE, month),
                month_column(ScheduleColumns.HOURS_TEMPLATE, month),
                month_column(ScheduleColumns.RATE_TEMPLATE, month),
            ])
            
        schedule_df = schedule_df[cols_order]
        schedule_df.to_excel(output_path, index=False, na_rep="")
        
        self._style_excel(output_path)
        
        return unassigned_warnings

    def _style_excel(self, filepath: str):
        wb = load_workbook(filepath)
        ws = wb.active
        
        pink_fill = PatternFill(start_color="F2DCDD", end_color="F2DCDD", fill_type="solid")
        
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            cell_val = str(row[1].value) 
            if 'סה"כ' in cell_val:
                for cell in row:
                    cell.fill = pink_fill
                    
        wb.save(filepath)